from pathlib import Path
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# Configuration
# ============================================================

INPUT_FILE = Path("paper_task_summary_tables_updated.xlsx")
OUTPUT_FILE = Path("paper_task_summary_tables_camera_ready_final.xlsx")

MAIN_SHEETS = ["PRED", "DESC", "KT", "ERS"]


# ============================================================
# General text utilities
# ============================================================

def normalize_whitespace(text):
    """
    Normalize spacing, line breaks, separators, and punctuation spacing.
    """
    if text is None:
        return text

    text = str(text)

    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)

    # Normalize common separators
    text = re.sub(r"\s*;\s*", "; ", text)
    text = re.sub(r"\s*,\s*", ", ", text)
    text = re.sub(r"\s*/\s*", " / ", text)

    # Avoid excessive spaces introduced by separator normalization
    text = re.sub(r"[ \t]+", " ", text)

    # Fix punctuation spacing
    text = re.sub(r"\s+\.", ".", text)
    text = re.sub(r"\s+\)", ")", text)
    text = re.sub(r"\(\s+", "(", text)
    text = re.sub(r"\s+:", ": ", text)
    text = re.sub(r":\s+", ": ", text)

    return text.strip()


def sentence_case_label(text):
    """
    Convert short labels to sentence case while preserving obvious acronyms.
    """
    if text is None:
        return text

    text = normalize_whitespace(text)

    if not text:
        return text

    # Preserve all-uppercase acronyms or short codes
    if text.isupper() and len(text) <= 6:
        return text

    return text[0].upper() + text[1:]


def remove_ellipsis_fragments(text):
    """
    Remove incomplete prose fragments that end in ellipses.

    Preserves useful numeric ellipses such as:
    - (0, 1, …, 10)
    - 1, 2, …, n
    """
    if text is None:
        return text

    text = str(text)

    # Normalize three dots to a single ellipsis
    text = text.replace("...", "…")

    lines = []

    for line in text.splitlines():
        stripped = line.strip()

        if not stripped:
            continue

        looks_like_numeric_range = bool(
            re.search(r"\d\s*,\s*\d\s*,?\s*…\s*,?\s*\d", stripped)
        )

        # Remove entire line if it is a prose fragment ending in ellipsis
        if stripped.endswith("…") and not looks_like_numeric_range:
            continue

        lines.append(line)

    text = "\n".join(lines)

    # Remove sentence-like prose fragments ending in ellipsis
    # while preserving numeric ranges.
    parts = re.split(r"(?<=[.;])\s+", text)
    cleaned_parts = []

    for part in parts:
        stripped = part.strip()

        if not stripped:
            continue

        looks_like_numeric_range = bool(
            re.search(r"\d\s*,\s*\d\s*,?\s*…\s*,?\s*\d", stripped)
        )

        if stripped.endswith("…") and not looks_like_numeric_range:
            continue

        cleaned_parts.append(stripped)

    return normalize_whitespace(" ".join(cleaned_parts))


# ============================================================
# First-pass standardization
# ============================================================

def first_pass_standardize_text(text):
    """
    First-pass cleanup:
    - typo corrections
    - consistent terminology
    - consistent separators
    - placeholder cleanup
    - general camera-ready wording
    """
    if text is None:
        return text

    text = str(text)

    replacements = {
        # Placeholder / obvious errors
        "110": "",
        "n.a.": "Not reported",
        "N.A.": "Not reported",
        "NA": "Not reported",
        "N/A": "Not reported",
        "nan": "Not reported",
        "None": "Not reported",

        # Typographical corrections
        "behaviour": "behavior",
        "Behaviour": "Behavior",
        "modelling": "modeling",
        "Modelling": "Modeling",
        "pre-processing": "preprocessing",
        "Pre-processing": "Preprocessing",
        "e-learning": "e-learning",
        "E-learning": "E-learning",

        # Common educational data mining terminology
        "drop out": "dropout",
        "Drop out": "Dropout",
        "drop-out": "dropout",
        "Drop-out": "Dropout",
        "at risk": "at-risk",
        "At risk": "At-risk",

        # Binary labels
        "correct / incorrect": "correct or incorrect",
        "Correct / incorrect": "Correct or incorrect",
        "right / wrong": "correct or incorrect",
        "Right / wrong": "Correct or incorrect",
        "pass / fail": "pass or fail",
        "Pass / fail": "Pass or fail",
        "yes / no": "yes or no",
        "Yes / no": "Yes or no",

        # Repeated task wording
        "Next Question is Correct": "Correctness of the next question",
        "Correct Answer in Next Question": "Correctness of the next question",
        "Performance on Next Question": "Performance on the next question",
        "next question is correct": "correctness of the next question",
        "next question correct": "correctness of the next question",

        # Casing consistency
        "Knowledge Tracing": "knowledge tracing",
        "Student Performance": "student performance",
        "Learning Analytics": "learning analytics",
        "Educational Data Mining": "educational data mining",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    text = remove_ellipsis_fragments(text)
    text = normalize_whitespace(text)

    # Convert empty strings created by placeholder removal
    if text == "":
        text = "Not reported"

    return text


# ============================================================
# Second-pass task-specification harmonization
# ============================================================

def canonical_task_specification(value, sheet_name):
    """
    Second-pass harmonization of near-duplicate Task Specification strings.

    The goal is not to erase meaningful differences, but to make highly similar
    labels use the same wording across rows.
    """
    if value is None:
        return value

    raw = normalize_whitespace(str(value))

    if not raw:
        return raw

    low = raw.lower()

    # --------------------------------------------------------
    # KT: Knowledge tracing task specifications
    # --------------------------------------------------------
    if sheet_name == "KT":

        if (
            "probability" in low
            and "next" in low
            and ("question" in low or "item" in low or "problem" in low)
            and ("correct" in low or "success" in low)
        ):
            return "Predict the probability that the student will answer the next question correctly"

        if (
            "next" in low
            and ("question" in low or "item" in low or "problem" in low)
            and ("correct" in low or "answer" in low)
        ):
            return "Predict whether the student will answer the next question correctly"

        if (
            "first attempt" in low
            and ("correct" in low or "success" in low)
        ):
            return "Predict whether the student will answer correctly on the first attempt"

        if (
            "next exercise" in low
            and ("correct" in low or "success" in low)
        ):
            return "Predict whether the student will solve the next exercise correctly"

        if (
            "exercise" in low
            and ("correct" in low or "success" in low)
        ):
            return "Predict whether the student will solve the exercise correctly"

        if (
            "programming" in low
            and ("correct" in low or "success" in low or "complete" in low)
        ):
            return "Predict whether the student will successfully complete the next programming task"

        if (
            "knowledge state" in low
            or "knowledge level" in low
            or "skill mastery" in low
            or "mastery" in low
        ):
            return "Estimate the student's knowledge state or skill mastery"

        if (
            "performance" in low
            and "next" in low
        ):
            return "Predict the student's performance on the next learning activity"

        if (
            "hint" in low
            or "help" in low
        ):
            return "Predict the student's need for hints or assistance"

    # --------------------------------------------------------
    # PRED: Predictive modeling task specifications
    # --------------------------------------------------------
    if sheet_name == "PRED":

        if "dropout" in low or "drop out" in low or "drop-out" in low:
            return "Predict student dropout risk"

        if "final grade" in low or "final mark" in low or "final score" in low:
            return "Predict the student's final grade"

        if "pass" in low and "fail" in low:
            return "Predict whether the student will pass or fail"

        if "at-risk" in low or "at risk" in low:
            return "Identify students at risk of poor academic performance"

        if (
            "academic performance" in low
            or "student performance" in low
            or ("performance" in low and "course" in low)
        ):
            return "Predict student academic performance"

        if "success" in low and "student" in low:
            return "Predict student academic success"

        if "grade" in low:
            return "Predict student grades"

    # --------------------------------------------------------
    # DESC: Descriptive / profiling task specifications
    # --------------------------------------------------------
    if sheet_name == "DESC":

        if "cluster" in low or "group" in low or "profile" in low:
            return "Describe or group students according to learning behavior"

        if "engagement" in low:
            return "Describe student engagement patterns"

        if "behavior" in low or "behaviour" in low:
            return "Describe student learning behavior"

        if "trajectory" in low or "pathway" in low or "sequence" in low:
            return "Describe student learning trajectories"

        if "pattern" in low:
            return "Describe patterns in student learning activity"

    # --------------------------------------------------------
    # ERS: Educational recommender systems task specifications
    # --------------------------------------------------------
    if sheet_name == "ERS":

        if "recommend" in low and "resource" in low:
            return "Recommend learning resources to students"

        if "recommend" in low and "course" in low:
            return "Recommend courses or learning paths to students"

        if "recommend" in low and "path" in low:
            return "Recommend courses or learning paths to students"

        if "recommend" in low and "activity" in low:
            return "Recommend learning activities to students"

        if "recommend" in low and "peer" in low:
            return "Recommend peers or collaborators to students"

        if "recommend" in low:
            return "Recommend personalized learning support to students"

    return first_pass_standardize_text(raw)


# ============================================================
# Header and worksheet helpers
# ============================================================

def find_header_row(ws, max_scan_rows=10):
    """
    Find the most likely header row by looking for known table headers.
    """
    expected_headers = {
        "task",
        "task specification",
        "context",
        "teaching method",
        "student performance definition",
        "target",
        "students",
        "courses",
        "data sources",
        "features",
        "features / variables",
        "features / representations",
        "moment of prediction",
        "preprocessing details",
        "models",
        "assessment strategy",
        "performance metric",
        "comments",
        "limitations",
    }

    best_row = 1
    best_score = 0

    for row in range(1, min(max_scan_rows, ws.max_row) + 1):
        values = []

        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value

            if value is not None:
                values.append(normalize_whitespace(str(value)).lower())

        score = sum(1 for value in values if value in expected_headers)

        if score > best_score:
            best_score = score
            best_row = row

    return best_row


def get_header_map(ws, header_row):
    """
    Return normalized header -> column index.
    """
    header_map = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value

        if value is None:
            continue

        key = normalize_whitespace(str(value)).lower()
        header_map[key] = col

    return header_map


def delete_feature_columns(ws, header_row):
    """
    Delete Features columns because they mostly repeat Data Sources.
    Applied in the second pass.
    """
    feature_headers = {
        "features",
        "feature",
        "features / variables",
        "features/variables",
        "features / representations",
        "features/representations",
        "features / data",
        "features/data",
    }

    cols_to_delete = []

    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value

        if value is None:
            continue

        key = normalize_whitespace(str(value)).lower()

        if key in feature_headers or key.startswith("features"):
            cols_to_delete.append(col)

    for col in sorted(cols_to_delete, reverse=True):
        ws.delete_cols(col)

    return len(cols_to_delete)


# ============================================================
# Column-specific cleanup
# ============================================================

def clean_target_column(value):
    """
    Standardize target labels while avoiding over-editing.
    """
    if value is None:
        return value

    text = first_pass_standardize_text(value)
    low = text.lower()

    target_replacements = {
        "correct answer": "Correct answer",
        "correctness": "Correctness",
        "student performance": "Student performance",
        "final grade": "Final grade",
        "dropout": "Dropout",
        "pass or fail": "Pass or fail",
        "knowledge state": "Knowledge state",
        "skill mastery": "Skill mastery",
    }

    for old, new in target_replacements.items():
        if low == old:
            return new

    return sentence_case_label(text)


def clean_data_sources(value):
    """
    Standardize data-source descriptions and separators.
    """
    if value is None:
        return value

    text = first_pass_standardize_text(value)

    replacements = {
        "LMS": "Learning management system",
        "VLE": "Virtual learning environment",
        "MOOC": "MOOC",
        "clickstream": "Clickstream data",
        "Clickstream": "Clickstream data",
        "logs": "Log data",
        "Logs": "Log data",
        "grades": "Grades",
        "Grades": "Grades",
        "demographic": "Demographic data",
        "Demographic": "Demographic data",
    }

    # Only replace exact short labels or semicolon-separated labels carefully
    parts = [p.strip() for p in re.split(r";|\n", text) if p.strip()]
    cleaned_parts = []

    for part in parts:
        key = part.strip()

        if key in replacements:
            cleaned_parts.append(replacements[key])
        else:
            cleaned_parts.append(sentence_case_label(key))

    if cleaned_parts:
        # Remove duplicates while preserving order
        seen = set()
        unique_parts = []

        for part in cleaned_parts:
            if part.lower() not in seen:
                unique_parts.append(part)
                seen.add(part.lower())

        return "; ".join(unique_parts)

    return text


def clean_models(value):
    """
    Standardize model names and separators without expanding every acronym.
    """
    if value is None:
        return value

    text = first_pass_standardize_text(value)

    replacements = {
        "random forest": "Random forest",
        "Random Forest": "Random forest",
        "decision tree": "Decision tree",
        "Decision Tree": "Decision tree",
        "logistic regression": "Logistic regression",
        "linear regression": "Linear regression",
        "support vector machine": "Support vector machine",
        "Support Vector Machine": "Support vector machine",
        "neural network": "Neural network",
        "Neural Network": "Neural network",
        "deep learning": "Deep learning",
        "bayesian": "Bayesian",
        "Bayesian": "Bayesian",
        "xgboost": "XGBoost",
        "Xgboost": "XGBoost",
        "lstm": "LSTM",
        "Lstm": "LSTM",
        "rnn": "RNN",
        "Rnn": "RNN",
        "dkt": "DKT",
        "Dkt": "DKT",
        "bkt": "BKT",
        "Bkt": "BKT",
    }

    for old, new in replacements.items():
        text = re.sub(rf"\b{re.escape(old)}\b", new, text)

    return normalize_whitespace(text)


def clean_performance_metric(value):
    """
    Standardize common metric labels.
    """
    if value is None:
        return value

    text = first_pass_standardize_text(value)

    replacements = {
        "accuracy": "Accuracy",
        "Accuracy": "Accuracy",
        "auc": "AUC",
        "Auc": "AUC",
        "roc": "ROC",
        "rmse": "RMSE",
        "Rmse": "RMSE",
        "mae": "MAE",
        "Mae": "MAE",
        "mse": "MSE",
        "Mse": "MSE",
        "f1": "F1-score",
        "F1": "F1-score",
        "f1-score-score": "F1-score",
        "precision": "Precision",
        "recall": "Recall",
        "Recall": "Recall",
    }

    for old, new in replacements.items():
        text = re.sub(rf"\b{re.escape(old)}\b", new, text)

    text = text.replace("F1-score-score", "F1-score")

    return normalize_whitespace(text)


def clean_cell_by_column(value, header, sheet_name):
    """
    Apply column-specific cleaning where appropriate.
    """
    if value is None:
        return value

    header = normalize_whitespace(str(header)).lower() if header else ""

    if header == "task specification":
        return canonical_task_specification(value, sheet_name)

    if header == "target":
        return clean_target_column(value)

    if header == "data sources":
        return clean_data_sources(value)

    if header == "models":
        return clean_models(value)

    if header == "performance metric":
        return clean_performance_metric(value)

    return first_pass_standardize_text(value)


# ============================================================
# Formatting
# ============================================================

def style_sheet(ws, header_row):
    """
    Apply paper-ready formatting:
    - bold filled headers
    - wrapped text
    - thin borders
    - frozen header row
    - sensible column widths
    """
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )

    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = border

    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
            cell.border = border

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        header_value = ws.cell(row=header_row, column=col).value
        header_text = str(header_value).lower() if header_value else ""

        if "reference" in header_text or "paper" in header_text or "study" in header_text:
            width = 24
        elif "task specification" in header_text:
            width = 38
        elif "data sources" in header_text:
            width = 36
        elif "models" in header_text:
            width = 32
        elif "assessment" in header_text:
            width = 30
        elif "metric" in header_text:
            width = 26
        elif "comments" in header_text or "limitations" in header_text:
            width = 40
        elif "preprocessing" in header_text:
            width = 36
        else:
            width = 24

        ws.column_dimensions[letter].width = width

    for row_idx in range(1, ws.max_row + 1):
        if row_idx == header_row:
            ws.row_dimensions[row_idx].height = 36
        else:
            ws.row_dimensions[row_idx].height = 30


# ============================================================
# Audit sheet
# ============================================================

def add_or_replace_audit_sheet(wb, audit_rows):
    """
    Add a compact audit sheet documenting the transformations.
    """
    if "audit" in wb.sheetnames:
        ws = wb["audit"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("audit")

    ws.append(["Sheet", "Pass", "Change", "Count / Notes"])

    for row in audit_rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="E2F0D9")
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 48


# ============================================================
# Main workbook processing
# ============================================================

def process_workbook(input_file, output_file):
    wb = load_workbook(input_file)

    audit_rows = []

    for sheet_name in MAIN_SHEETS:
        if sheet_name not in wb.sheetnames:
            audit_rows.append([
                sheet_name,
                "Both",
                "Skipped",
                "Sheet not found",
            ])
            continue

        ws = wb[sheet_name]

        header_row = find_header_row(ws)
        header_map = get_header_map(ws, header_row)

        # ----------------------------------------------------
        # First pass: general text standardization
        # ----------------------------------------------------
        first_pass_changed_cells = 0

        for row in range(header_row + 1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)

                if not isinstance(cell.value, str):
                    continue

                original = cell.value
                updated = first_pass_standardize_text(original)

                if updated != original:
                    cell.value = updated
                    first_pass_changed_cells += 1

        audit_rows.append([
            sheet_name,
            "First pass",
            "General text cleanup, terminology standardization, typo correction, separator normalization, and ellipsis-fragment removal",
            first_pass_changed_cells,
        ])

        # ----------------------------------------------------
        # Second pass: delete repeated Features columns
        # ----------------------------------------------------
        deleted_feature_columns = 0

        if sheet_name in {"PRED", "DESC", "KT"}:
            deleted_feature_columns = delete_feature_columns(ws, header_row)

        audit_rows.append([
            sheet_name,
            "Second pass",
            "Deleted Features column(s) because they mostly repeated Data Sources",
            deleted_feature_columns,
        ])

        # Rebuild header map after deletion
        header_map = get_header_map(ws, header_row)

        # ----------------------------------------------------
        # Second pass: column-aware cleanup and harmonization
        # ----------------------------------------------------
        second_pass_changed_cells = 0
        task_spec_changed_cells = 0

        for row in range(header_row + 1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)

                if not isinstance(cell.value, str):
                    continue

                header = ws.cell(row=header_row, column=col).value
                header_key = normalize_whitespace(str(header)).lower() if header else ""

                original = cell.value
                updated = clean_cell_by_column(original, header, sheet_name)

                if updated != original:
                    cell.value = updated
                    second_pass_changed_cells += 1

                    if header_key == "task specification":
                        task_spec_changed_cells += 1

        audit_rows.append([
            sheet_name,
            "Second pass",
            "Column-aware cleanup and harmonization of repeated wording",
            second_pass_changed_cells,
        ])

        audit_rows.append([
            sheet_name,
            "Second pass",
            "Standardized Task Specification wording",
            task_spec_changed_cells,
        ])

        # ----------------------------------------------------
        # Formatting
        # ----------------------------------------------------
        style_sheet(ws, header_row)

        audit_rows.append([
            sheet_name,
            "Both",
            "Applied camera-ready formatting",
            "Wrapped text, frozen headers, compact widths, borders, and header styling",
        ])

    add_or_replace_audit_sheet(wb, audit_rows)

    wb.save(output_file)

    return output_file


if __name__ == "__main__":
    saved_file = process_workbook(INPUT_FILE, OUTPUT_FILE)
    print(f"Saved standardized workbook to: {saved_file}")